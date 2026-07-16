#!/usr/bin/env python3
"""
Monthly staff shift scheduler  (configurable engine).

Generates a valid roster for ANY month using a CP-SAT constraint solver
(Google OR-Tools), then independently validates every scheduling rule and
exports a color-coded Excel workbook (roster sheet + summary/validation sheet).

Why CP-SAT: every hard rule is modelled as an explicit constraint, so the
solver either returns a schedule that satisfies *all* of them simultaneously
or proves the rule set is infeasible. We never relax a hard rule to "make it
fit". After solving, a separate validator re-derives every rule from the
finished grid (it does not trust the solver) and reports PASS/FAIL.

All configuration lives in one `ScheduleSettings` object so the same engine
backs both the CLI (`python scheduler.py`) and the Streamlit app (`app.py`):
nothing about the month, the staff, or the rule values is hard-coded.

POLICY (this revision):
  * Every employee works EXACTLY `shifts_per_employee` shifts for the month.
  * Leave (vacation) date ranges hard-block those days -- shown as "V" in the
    grid -- and prorate that employee's exact total to max(0, shifts - leave
    days). Run rules apply within each free stretch between leave blocks.
  * Nights are covered under one of two selectable rotation modes:
      - "fixed_team": a dedicated night team of ANY size N works nights ONLY and
        is the only group eligible for nights. Coverage is the band
        [night_min, night_max]; overlap above the floor appears only on the days
        the exact arithmetic forces it, never as a free-for-all.
      - "rotate": there is no dedicated team -- every employee is night-eligible
        and nights rotate across the whole roster. The night load is balanced as
        evenly as possible (a soft fairness goal), not pinned to a fixed total.
  * No Night->Day next-calendar-day transition for anyone.
  * Day-shift runs are capped at `max_consec_work`; night runs at
    `max_consec_night`; off runs floored at `min_consec_off`, capped at
    `max_consec_off`; work runs floored at `min_consec_work`.
  * FAIRNESS is the top priority and is expressed as four soft, equally-weighted
    objective goals -- equal totals, balanced night load, balanced weekends, and
    balanced "undesirable runs" (forced overlaps + max-length streaks) -- each of
    which the independent validator re-derives and reports as PASS/FAIL plus the
    measured spread (max - min). Fairness goals are SOFT: they never block a
    solve, and never override a hard rule.

Run:  ./.venv/bin/python scheduler.py
Out:  schedule.xlsx  (+ a summary printed to stdout)
"""

from __future__ import annotations

import calendar
import datetime
import sys
from dataclasses import dataclass, field
from io import BytesIO

from ortools.sat.python import cp_model
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Shift codes used throughout.
DAY, NIGHT, OFF = "D", "N", "O"
# Leave (vacation): configured per ScheduleSettings.leave, hard-blocked in the
# model, and labelled "V" in the grid so both views render it from one artifact.
LEAVE = "V"
# AM staff work a fixed weekly pattern and live entirely outside the solver (see
# am_rows): they are appended to the roster, never scheduled, counted, or validated.
AM = "AM"

# Rotation modes (how nights are covered).
FIXED_TEAM, ROTATE = "fixed_team", "rotate"

# Calendar weekday names, indexed Sun..Sat to match the roster header style.
WEEKDAY_NAMES = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]


# ---------------------------------------------------------------------------
# SETTINGS  (the single source of truth the UI populates)
# ---------------------------------------------------------------------------

@dataclass
class ScheduleSettings:
    # --- Calendar -----------------------------------------------------------
    year: int = 2026
    month: int = 8                                  # 1-12

    # --- Staff --------------------------------------------------------------
    employees: list[str] = field(
        default_factory=lambda: [f"Employee {i}" for i in range(1, 8)])

    # --- Night model --------------------------------------------------------
    # "fixed_team": night_team (ANY size) works nights only and is the only
    #               group eligible for nights.
    # "rotate":     no dedicated team -- everyone is night-eligible.
    rotation_mode: str = FIXED_TEAM
    night_team: list[str] = field(
        default_factory=lambda: ["Employee 6", "Employee 7"])
    night_team_nights_only: bool = True             # fixed_team: no day shifts

    # --- AM shift -----------------------------------------------------------
    # AM staff work a FIXED weekly pattern (am_days) and are entirely OUTSIDE the
    # solver: never scheduled, counted for coverage, or validated -- just appended
    # to the roster as extra rows. am_team defaults empty (the feature is opt-in);
    # am_days defaults to the Saudi workweek (Sun-Thu) and is configurable.
    am_team: list[str] = field(default_factory=list)
    am_days: tuple = ("Sun", "Mon", "Tue", "Wed", "Thu")

    # --- Leave (vacation) ----------------------------------------------------
    # One row per leave range: (employee name, first day, last day), dates
    # inclusive, within the chosen month. Multiple rows per employee allowed.
    # Empty by default: with no leave the model is byte-identical to the
    # pre-feature engine. AM staff cannot take rostered leave (preflight rejects).
    leave: list[tuple[str, datetime.date, datetime.date]] = field(default_factory=list)

    # --- Staffing bands -----------------------------------------------------
    day_min: int = 2                                # day staff required per day
    day_max: int = 4
    night_min: int = 1                              # >= 1 night every day
    night_max: int = 2                              # allow up to 2 overlapping

    # --- Workload -----------------------------------------------------------
    shifts_per_employee: int = 16                   # EXACT monthly total
    hours_per_shift: int = 12                       # one shift = 12 h (Hours col)

    # --- Run lengths --------------------------------------------------------
    max_consec_work: int = 4                        # combined working (day+night) run cap
    max_consec_night: int = 3                       # night run cap (priority)
    min_consec_work: int = 2                        # no isolated single work day
    min_consec_off: int = 2                         # no isolated single off day
    max_consec_off: int = 4                         # no long idle block

    # --- Fairness -----------------------------------------------------------
    weekend_days: tuple = ("Fri", "Sat")            # the "weekend" for fairness

    # HARD toggle (NOT one of the four soft fairness goals): when True, every
    # employee's full (Fri, Sat) weekends strictly alternate off/on/off/... --
    # of every two consecutive full weekends, exactly one is fully off. The
    # solver picks each person's phase; months with < 2 full weekends make it
    # vacuous. Unlike w_fair_* it is enforced as a constraint and re-derived as a
    # PASS/FAIL hard rule, never a spread. Setting it False leaves the model
    # byte-identical to the pre-feature engine, so it defaults OFF (opt-in):
    # pinning WHICH weekends each person is off can over-constrain a tight month
    # to INFEASIBLE (e.g. the default 7-staff / 2-night-team config in a 28-day
    # February), and it structurally fights the F3 "even Fri/Sat duty" sub-goal
    # (a small night pool over an odd number of full weekends must split phases,
    # forcing a weekend spread > fair_tol_weekend). Both are honestly surfaced --
    # preflight/relaxation_hints for the former, an honest soft F3 FAIL (every
    # HARD rule still holds) for the latter -- so neither belongs in the
    # out-of-box experience.
    alternating_weekends: bool = False

    # Tolerances: a fairness goal PASSES when its measured spread (max - min)
    # is within tolerance. Loose by default so common months stay green; tighten
    # from the UI to demand stricter fairness. Totals are pinned exact (spread 0).
    fair_tol_total: int = 0
    fair_tol_night: int = 1
    fair_tol_weekend: int = 1
    fair_tol_runs: int = 2

    # --- Objective weights (all minimized; equal => the four goals matter
    #     equally). Raising one weight prioritises that goal in trade-offs. -----
    w_fair_total: int = 100                         # equal total shifts
    w_fair_night: int = 100                         # balanced night load
    w_fair_weekend: int = 100                       # balanced weekends
    w_fair_runs: int = 100                          # balanced undesirable runs

    # Determinism: the solver is bounded by DETERMINISTIC time (reproducible
    # work units, machine-independent) so the same inputs always give the same
    # roster -- even when a hard fairness objective is not proven optimal in
    # time. Strict reproducibility relies on THIS limit binding first; the
    # wall-clock `solver_time_limit` is only a safety net, sized so it does not
    # bind before the deterministic budget on normal hardware. Fixed-team months
    # solve in well under a second; the larger budget is headroom for the harder
    # rotate mode, whose soft weekend goal needs more search to converge.
    solver_det_time_limit: float = 24.0
    solver_time_limit: float = 180.0

    # --- Derived calendar facts --------------------------------------------
    @property
    def days(self) -> int:
        return calendar.monthrange(self.year, self.month)[1]

    @property
    def start_weekday(self) -> str:
        wd = datetime.date(self.year, self.month, 1).weekday()   # Mon=0..Sun=6
        return WEEKDAY_NAMES[(wd + 1) % 7]

    @property
    def month_label(self) -> str:
        return f"{calendar.month_name[self.month]} {self.year}"

    @property
    def is_rotate(self) -> bool:
        return self.rotation_mode == ROTATE

    @property
    def day_team(self) -> list[str]:
        """Employees who may work day shifts (everyone, in rotate mode)."""
        if self.is_rotate or not self.night_team_nights_only:
            return list(self.employees)
        return [e for e in self.employees if e not in self.night_team]


# ---------------------------------------------------------------------------
# CALENDAR / ROLE HELPERS
# ---------------------------------------------------------------------------

def weekday_of(settings: ScheduleSettings, day_index: int) -> str:
    start = WEEKDAY_NAMES.index(settings.start_weekday)
    return WEEKDAY_NAMES[(start + day_index) % 7]


def weekend_day_indices(settings: ScheduleSettings) -> list[int]:
    return [d for d in range(settings.days)
            if weekday_of(settings, d) in settings.weekend_days]


def weekend_pair_indices(settings: ScheduleSettings) -> list[tuple[int, int]]:
    """(Fri, Sat) consecutive index pairs that make a full weekend."""
    return [(d, d + 1) for d in range(settings.days - 1)
            if weekday_of(settings, d) == "Fri" and weekday_of(settings, d + 1) == "Sat"]


def night_eligible_indices(settings: ScheduleSettings) -> list[int]:
    """Employee indices allowed to work nights (everyone, in rotate mode)."""
    if settings.is_rotate:
        return list(range(len(settings.employees)))
    return [settings.employees.index(name) for name in settings.night_team
            if name in settings.employees]


def day_capable_indices(settings: ScheduleSettings) -> list[int]:
    """Employee indices allowed to work day shifts.

    In fixed_team + nights-only mode the night team never works days; otherwise
    (rotate, or a fixed team that may also work days) everyone is day-capable.
    """
    n = len(settings.employees)
    if not settings.is_rotate and settings.night_team_nights_only:
        elig = set(night_eligible_indices(settings))
        return [i for i in range(n) if i not in elig]
    return list(range(n))


def is_night_member(settings: ScheduleSettings, e: int) -> bool:
    """True iff employee index `e` belongs to a dedicated nights-only team."""
    return (not settings.is_rotate
            and settings.employees[e] in settings.night_team)


def leave_day_sets(settings: ScheduleSettings) -> list[set[int]]:
    """Per-employee 0-based day indices on leave (union of the configured ranges).

    Overlapping ranges collapse; each range is clamped to the month; unknown
    names are skipped. Clamping keeps this total (defense in depth) --
    preflight() is the authority that rejects out-of-month or malformed ranges.
    date -> index: (date - first_of_month).days.
    """
    S = settings
    first = datetime.date(S.year, S.month, 1)
    out: list[set[int]] = [set() for _ in S.employees]
    index = {name: i for i, name in enumerate(S.employees)}
    for name, d0, d1 in S.leave:
        if name not in index:
            continue
        if isinstance(d0, datetime.datetime):
            d0 = d0.date()
        if isinstance(d1, datetime.datetime):
            d1 = d1.date()
        a, b = (d0 - first).days, (d1 - first).days
        out[index[name]].update(range(max(0, a), min(S.days - 1, b) + 1))
    return out


def shift_targets(settings: ScheduleSettings) -> list[int]:
    """Per-employee exact monthly total: shifts_per_employee prorated by leave."""
    lv = leave_day_sets(settings)
    return [max(0, settings.shifts_per_employee - len(lv[e]))
            for e in range(len(settings.employees))]


def available_segments(settings: ScheduleSettings) -> list[list[tuple[int, int]]]:
    """Per-employee maximal (start, end) inclusive runs of non-leave day indices.

    The run-length rules apply WITHIN each segment (leave days belong to no
    run); with no leave every employee has the single segment (0, days-1), which
    makes the per-segment encodings collapse to the pre-feature constraints.
    """
    S = settings
    lv = leave_day_sets(S)
    segs_all: list[list[tuple[int, int]]] = []
    for e in range(len(S.employees)):
        segs, start = [], None
        for d in range(S.days):
            if d in lv[e]:
                if start is not None:
                    segs.append((start, d - 1))
                    start = None
            elif start is None:
                start = d
        if start is not None:
            segs.append((start, S.days - 1))
        segs_all.append(segs)
    return segs_all


def expected_double_nights(settings: ScheduleSettings) -> int:
    """Forced night-overlap (extra nights above the floor), given exact totals.

    A nights-only team of ANY size N, each working an exact target (their
    `shifts_per_employee` prorated by leave), places a fixed summed-target
    number of night-shifts. With >= night_min covered every day, the
    surplus above the floor -- max(0, total_nights - night_min*days) -- must land
    as overlap (extra nights stacked onto some days). This is the minimum the
    arithmetic forces; nothing more. The validator checks the realised overlap
    equals it. For the default 2-person team on night_min=1 this is the classic
    "(32 - days) double-night days". In 'rotate' mode the night total is not
    pinned, so there is no forced overlap and this returns 0.
    """
    if settings.is_rotate or not settings.night_team_nights_only:
        return 0
    tgt = shift_targets(settings)
    total_nights = sum(tgt[e] for e in night_eligible_indices(settings))
    return max(0, total_nights - settings.night_min * settings.days)


def roster_caption(settings: ScheduleSettings) -> str:
    """One-line description of the roster, shared by Excel and the CLI."""
    S = settings
    head = f"Roster: {S.month_label} ({S.days} days)."
    n_leave = sum(1 for lv in leave_day_sets(S) if lv)
    tail = f"; {n_leave} staff on leave." if n_leave else "."
    if S.is_rotate:
        return f"{head} All staff rotate nights{tail}"
    names = ", ".join(S.night_team) if S.night_team else "(none)"
    only = " (nights only)" if S.night_team_nights_only else ""
    return f"{head} Night team: {names}{only}{tail}"


# ---------------------------------------------------------------------------
# AM SHIFT  (fixed weekly-pattern staff, entirely outside the solver)
# ---------------------------------------------------------------------------

def am_rows(settings: ScheduleSettings) -> list[list[str]]:
    """One roster row per AM-shift member, index-aligned with settings.am_team.

    AM staff live entirely OUTSIDE the solver: they are never modelled as
    constraints, never counted in coverage or fairness, and never validated --
    they are simply appended as extra rows to the finished roster (the Streamlit
    grid and the Excel sheet). Cell d is AM when that calendar day's weekday name
    is in am_days, else OFF. Weekday names are derived exactly the way the rest of
    the engine derives them (weekday_of), so AM rows share the roster's calendar.
    """
    return [[AM if weekday_of(settings, d) in settings.am_days else OFF
             for d in range(settings.days)]
            for _ in settings.am_team]


# ---------------------------------------------------------------------------
# FEASIBILITY PRE-FLIGHT  (catches a bad rule set before the solver, and -- per
# the UI choice -- pairs each problem with a concrete, non-binding suggestion)
# ---------------------------------------------------------------------------

@dataclass
class Problem:
    message: str
    suggestion: str


def _ceil(a: int, b: int) -> int:
    return -(-a // b)


def _segment_work_counts(length: int, max_work: int, max_off: int,
                         min_run: int) -> set[int]:
    """Feasible work totals for ONE free stretch of `length` days.

    Exact enumeration of the run rules inside a segment: work and off blocks
    alternate, work blocks in [min_run, max_work], off blocks in
    [min_run, max_off]. A stretch too short for any legal block (a singleton
    between leave/month edges) is forced fully off -> {0}, mirroring the
    solver's singleton rule and the validator's whole-segment exemption.
    """
    if length < min_run:
        return {0}
    # states[(filled, last_block_kind)] -> set of reachable work totals.
    states: dict[tuple[int, str | None], set[int]] = {(0, None): {0}}
    for pos in range(length):
        for last in (None, "w", "o"):
            reached = states.get((pos, last))
            if not reached:
                continue
            if last != "w":
                for blk in range(min_run, max_work + 1):
                    if pos + blk <= length:
                        states.setdefault((pos + blk, "w"), set()).update(
                            c + blk for c in reached)
            if last != "o":
                for blk in range(min_run, max_off + 1):
                    if pos + blk <= length:
                        states.setdefault((pos + blk, "o"), set()).update(reached)
    return states.get((length, "w"), set()) | states.get((length, "o"), set())


def preflight(settings: ScheduleSettings) -> list[Problem]:
    S = settings
    problems: list[Problem] = []
    days = S.days
    n_emp = len(S.employees)
    w = S.shifts_per_employee

    # --- AM shift (fixed-pattern staff, entirely outside the solver) --------
    # AM staff are appended as extra roster rows, never modelled, so their only
    # failure modes are naming / config mistakes. Cheap, mode-independent checks,
    # each paired with a concrete suggestion (like every other Problem). They run
    # before the empty-roster early return so AM mistakes are never masked.
    am_overlap = [name for name in S.am_team if name in S.employees]
    if am_overlap:
        problems.append(Problem(
            f"AM staff also appear in the scheduled employee list: {', '.join(am_overlap)}.",
            "Remove these names from either the AM list or the employee list -- AM "
            "staff are rostered separately, outside the solver."))
    am_seen, am_dupes = set(), []
    for name in S.am_team:
        if name in am_seen and name not in am_dupes:
            am_dupes.append(name)
        am_seen.add(name)
    if am_dupes:
        problems.append(Problem(
            f"Duplicate name(s) in the AM list: {', '.join(am_dupes)}.",
            "Give each AM staff member a single, unique row."))
    bad_am_days = [d for d in S.am_days if d not in WEEKDAY_NAMES]
    if bad_am_days:
        problems.append(Problem(
            f"AM working days include invalid weekday name(s): {', '.join(bad_am_days)}.",
            f"Use weekday names from {', '.join(WEEKDAY_NAMES)}."))
    if S.am_team and not S.am_days:
        problems.append(Problem(
            "AM staff are configured but no AM working days are set.",
            "Pick at least one AM working day, or clear the AM staff list."))

    # --- Basic shape (mode-independent) ------------------------------------
    if n_emp == 0:
        problems.append(Problem("No employees configured.",
                                "Add at least 3-4 employees."))
        return problems

    if S.min_consec_work != 2 or S.min_consec_off != 2:
        problems.append(Problem(
            "Minimum consecutive work/off must be 2 (the model encodes exactly 2).",
            "Keep 'min consecutive work' and 'min consecutive off' at 2."))
    if w > days:
        problems.append(Problem(
            f"{w} shifts/employee cannot fit in a {days}-day month.",
            f"Lower shifts/employee to at most {days}, or choose a longer month."))
        return problems
    if S.day_min + S.night_min > n_emp:
        problems.append(Problem(
            f"Each day needs at least {S.day_min} day + {S.night_min} night = "
            f"{S.day_min + S.night_min} people, but only {n_emp} employees exist.",
            f"Add staff, or lower the daily minimums to total <= {n_emp}."))

    # The run-length pattern check is shared by both modes.
    def pattern_problem(label, work_days, max_work, max_off):
        o = days - work_days
        bw_lo, bw_hi = _ceil(work_days, max_work), work_days // S.min_consec_work
        bo_lo = _ceil(o, max_off) if o > 0 else 0
        bo_hi = o // S.min_consec_off if o > 0 else 0
        if bw_lo > bw_hi:
            return Problem(
                f"{label}: {work_days} work days can't be split into blocks of "
                f"{S.min_consec_work}-{max_work}.",
                f"Raise the max {label.split()[0].lower()} run, or change shifts/employee.")
        if o > 0 and bo_lo > bo_hi:
            return Problem(
                f"{label}: {o} off days can't be split into blocks of "
                f"{S.min_consec_off}-{max_off}.",
                "Raise max consecutive off, or change shifts/employee.")
        if not any(abs(bw - bo) <= 1
                   for bw in range(bw_lo, bw_hi + 1)
                   for bo in range(bo_lo, bo_hi + 1)):
            return Problem(
                f"{label}: cannot alternate {work_days} work and {o} off days under "
                "the run-length rules.",
                "Widen a run-length cap, or change shifts/employee.")
        return None

    # --- Alternating weekends (HARD toggle; applies to BOTH modes) ----------
    # Strict off/on alternation puts every employee off on ~half the full
    # weekends, so the solver can split a pool of P into phases of at most
    # ceil(P/2) and floor(P/2); whichever is on the "worst" weekend leaves only
    # floor(P/2) available. Necessary (NOT sufficient): that worst-case half must
    # still meet the daily floor. Disjoint pools (nights-only team) are checked
    # per pool; overlapping pools (rotate, or a team that also works days) share
    # one floor. Only meaningful with >= 2 full weekends (else the rule is vacuous).
    if S.alternating_weekends and len(weekend_pair_indices(S)) >= 2:
        pools_disjoint = (not S.is_rotate) and S.night_team_nights_only
        if pools_disjoint:
            n_day = len(day_capable_indices(S))
            n_ni = len(night_eligible_indices(S))
            if n_day // 2 < S.day_min:
                problems.append(Problem(
                    f"Alternating weekends: at most {n_day // 2} of {n_day} day-capable "
                    f"staff can be on any weekend, below the {S.day_min} day-staff minimum.",
                    "Turn off alternating weekends, lower min day staff, or add day staff."))
            if n_ni // 2 < S.night_min:
                problems.append(Problem(
                    f"Alternating weekends: at most {n_ni // 2} of {n_ni} night-eligible "
                    f"staff can be on any weekend, below the {S.night_min} nights/day minimum.",
                    "Turn off alternating weekends, lower min nights/day, or add night staff."))
        else:
            if n_emp // 2 < S.day_min + S.night_min:
                problems.append(Problem(
                    f"Alternating weekends: at most {n_emp // 2} of {n_emp} staff can be on "
                    f"any weekend, below the {S.day_min + S.night_min} day+night daily floor.",
                    "Turn off alternating weekends, lower the daily minimums, or add staff."))

    # --- Leave (vacation): input validation, then exact necessities ---------
    # Input first: preflight is the authority on malformed ranges (the helpers
    # clamp defensively but never reject). AM staff cannot take rostered leave.
    if S.leave:
        first_day = datetime.date(S.year, S.month, 1)
        last_day = datetime.date(S.year, S.month, days)
        for name, d0, d1 in S.leave:
            if isinstance(d0, datetime.datetime):
                d0 = d0.date()
            if isinstance(d1, datetime.datetime):
                d1 = d1.date()
            if name in S.am_team:
                problems.append(Problem(
                    f"Leave entry names {name}, who is AM staff.",
                    "AM staff cannot take rostered leave in this version -- remove "
                    "the row, or move them into the scheduled employee list first."))
            elif name not in S.employees:
                problems.append(Problem(
                    f"Leave entry names {name}, who is not in the employee list.",
                    "Pick names from the employee list; AM staff cannot take "
                    "rostered leave in this version."))
            if d0 > d1:
                problems.append(Problem(
                    f"Leave for {name} starts after it ends ({d0} > {d1}).",
                    "Swap the dates so the range runs start to end."))
            elif d0 < first_day or d1 > last_day:
                problems.append(Problem(
                    f"Leave for {name} runs outside {S.month_label}.",
                    "Keep each range inside the month; split cross-month leave "
                    "per month."))

    lv = leave_day_sets(S)
    tgt = shift_targets(S)
    has_leave = any(lv)
    if has_leave:
        # Every check below is an exact necessity -- never a heuristic that
        # could false-block a feasible month (preflight hard-stops the app).
        def avail(pool, d):
            """Members of `pool` free on day d (not on leave, target > 0)."""
            return sum(1 for e in pool if tgt[e] > 0 and d not in lv[e])

        elig_pool = night_eligible_indices(S)
        day_pool = day_capable_indices(S)
        pools_disjoint = (not S.is_rotate) and S.night_team_nights_only

        # Per-day availability floors. Nights are workable only by the eligible
        # pool in EVERY mode, so its floor is always necessary; the day side is
        # per-pool when the pools are disjoint (nights-only team), else shared.
        def floor_problem(pool, floor, what):
            short = [d for d in range(days) if avail(pool, d) < floor]
            if short:
                dates = ", ".join(f"D{d + 1}" for d in short[:4]) + (
                    "..." if len(short) > 4 else "")
                problems.append(Problem(
                    f"Leave leaves fewer than {floor} {what} available on {dates}.",
                    "Stagger the overlapping leave or lower the daily minimums."))

        floor_problem(elig_pool, S.night_min, "night-eligible staff")
        if pools_disjoint:
            floor_problem(day_pool, S.day_min, "day staff")
        else:
            floor_problem(list(range(n_emp)), S.day_min + S.night_min,
                          "staff (day+night floor)")

        # Pinned-crew stretches: on days where a pool's availability EQUALS its
        # floor, every free member must work; anyone free through such a stretch
        # for longer than the pool's run cap is forced over the cap -- provably
        # infeasible even though the capacity sums and per-day floors pass.
        def pinned_problem(pool, floor, cap, what, unit="days"):
            if floor <= 0:
                return
            d = 0
            while d < days:
                if avail(pool, d) != floor:
                    d += 1
                    continue
                a = d
                while d < days and avail(pool, d) == floor:
                    d += 1
                b = d - 1                       # maximal pinned run [a, b]
                for e in pool:
                    if tgt[e] <= 0:
                        continue
                    run = best = 0
                    for x in range(a, b + 1):
                        run = 0 if x in lv[e] else run + 1
                        best = max(best, run)
                    if best > cap:
                        problems.append(Problem(
                            f"Leave pins the {what} crew at its minimum on "
                            f"D{a + 1}-D{b + 1}, forcing {S.employees[e]} to work "
                            f"more than {cap} {unit} straight.",
                            "Shorten or split the leave, or add cover for those days."))

        pinned_problem(elig_pool, S.night_min, S.max_consec_night, "night",
                       unit="nights")
        if pools_disjoint:
            pinned_problem(day_pool, S.day_min, S.max_consec_work, "day")
        else:
            pinned_problem(list(range(n_emp)), S.day_min + S.night_min,
                           S.max_consec_work, "day+night")

        # Per-employee pattern guard for leave-takers: an exact per-segment
        # enumeration (the contiguous pattern check below would falsely block
        # feasible split months). Each free stretch admits a set of work totals;
        # the prorated target must be reachable as a sum across stretches.
        segs_all = available_segments(S)
        for e in range(n_emp):
            if not lv[e] or tgt[e] == 0:
                continue
            cap = (S.max_consec_night
                   if not S.is_rotate and S.night_team_nights_only
                   and S.employees[e] in S.night_team
                   else S.max_consec_work)
            reach = {0}
            for (a, b) in segs_all[e]:
                counts = _segment_work_counts(b - a + 1, cap, S.max_consec_off,
                                              S.min_consec_work)
                reach = {r + c for r in reach for c in counts}
            if tgt[e] not in reach:
                problems.append(Problem(
                    f"{S.employees[e]}'s leave splits the month into stretches "
                    f"that cannot hold exactly {tgt[e]} shifts under the run rules.",
                    "Shift or split the leave, or change shifts/employee -- the "
                    "free stretches can't fit that many shifts in legal blocks."))

    # --- Mode B: everyone rotates nights -----------------------------------
    if S.is_rotate:
        total_shifts = sum(tgt)             # == w * n_emp when no leave
        floor_need = (S.day_min + S.night_min) * days
        ceil_cap = (S.day_max + S.night_max) * days
        if total_shifts < floor_need:
            problems.append(Problem(
                f"{total_shifts} total shifts can't cover the daily floor of "
                f"{S.day_min + S.night_min}/day ({floor_need} needed).",
                "Add staff, raise shifts/employee, or lower the daily minimums."))
        if total_shifts > ceil_cap:
            problems.append(Problem(
                f"{total_shifts} total shifts exceed the {ceil_cap} the daily maxima "
                f"allow (<= {S.day_max + S.night_max}/day).",
                "Raise max day/night staffing, remove staff, or lower shifts/employee."))
        # The contiguous pattern check speaks for leave-free employees only;
        # leave-takers were covered by the exact per-segment guard above.
        if not has_leave or any(not lv[i] for i in range(n_emp)):
            p = pattern_problem("Roster pattern", w, S.max_consec_work, S.max_consec_off)
            if p:
                problems.append(p)
        return problems

    # --- Mode A: fixed night team ------------------------------------------
    n_night = len(S.night_team)
    if not set(S.night_team).issubset(set(S.employees)):
        problems.append(Problem(
            "Night team includes someone who is not in the employee list.",
            "Choose night-team members from the current employee list."))
    if n_night < 1:
        problems.append(Problem(
            "No night-team members selected.",
            "Select at least 1 employee for the night team (any size)."))

    # Night coverage capacity (nights-only team). Supply is the summed
    # per-member TARGET (prorated by leave) -- identical to w * n_night when no
    # leave; a name not in the employee list still counts w so the message
    # matches the pre-leave arithmetic (the subset check above already fires).
    team_idx = [S.employees.index(nm) for nm in S.night_team if nm in S.employees]
    total_nights = (sum(tgt[i] for i in team_idx) + w * (n_night - len(team_idx))
                    if S.night_team_nights_only else None)
    if S.night_team_nights_only and n_night >= 1:
        if total_nights < S.night_min * days:
            if any(lv[i] for i in team_idx):
                sugg = ("Move a day-team member into the night team for this "
                        "month, or schedule the leave in a month with different "
                        "cover.")
            else:
                sugg = "Lower min nights/day, or give the night team more shifts."
            problems.append(Problem(
                f"Night team supplies {total_nights} nights, but >= {S.night_min}/day "
                f"needs {S.night_min * days}.",
                sugg))
        if total_nights > S.night_max * days:
            need = _ceil(total_nights, days)
            problems.append(Problem(
                f"Night team must place {total_nights} nights, but <= {S.night_max}/day "
                f"only allows {S.night_max * days}.",
                f"Raise max nights/day to {need}, add a night-team member, or shrink "
                f"the night-team workload."))

    # Day coverage capacity (summed targets; == w * n_emp - nights when no leave).
    day_shifts = sum(tgt) - (total_nights if total_nights is not None else 0)
    if day_shifts < S.day_min * days:
        problems.append(Problem(
            f"Only {day_shifts} day-shifts are available, but >= {S.day_min}/day "
            f"needs {S.day_min * days}.",
            f"Add day-team staff, or lower min day staffing to {day_shifts // days}."))
    if day_shifts > S.day_max * days:
        need = _ceil(day_shifts, days)
        problems.append(Problem(
            f"{day_shifts} day-shifts must be placed, but <= {S.day_max}/day only "
            f"allows {S.day_max * days}.",
            f"Raise max day staffing to {need}, or remove a day-team member."))

    # Per-employee run-length pattern feasibility. The contiguous check speaks
    # for leave-free pool members only; leave-takers were covered by the exact
    # per-segment guard above.
    if not has_leave or any(not lv[i] for i in day_capable_indices(S)):
        p = pattern_problem("Day-team pattern", w, S.max_consec_work, S.max_consec_off)
        if p:
            problems.append(p)
    if S.night_team_nights_only and n_night >= 1:
        if not has_leave or any(not lv[i] for i in team_idx):
            p = pattern_problem("Night-team pattern", w, S.max_consec_night, S.max_consec_off)
            if p:
                problems.append(p)

    return problems


def relaxation_hints(settings: ScheduleSettings) -> list[str]:
    """Best-effort suggestions when the model passes preflight but the solver
    still proves no schedule exists (the binding conflict is a combination, not a
    single arithmetic check)."""
    S = settings
    hints = [
        f"Widen the day-staffing band (e.g. max day staff {S.day_max} -> {S.day_max + 1}).",
        f"Raise max consecutive off ({S.max_consec_off} -> {S.max_consec_off + 1}) "
        f"or max consecutive working run ({S.max_consec_work} -> {S.max_consec_work + 1}).",
        f"Allow more night overlap (max nights/day {S.night_max} -> {S.night_max + 1}).",
        "Reduce shifts/employee by 1 to loosen the packing.",
    ]
    if S.alternating_weekends:
        hints.append(
            "Turn off alternating weekends -- forcing every employee off on every "
            "other full weekend can over-constrain a tight month.")
    if S.leave:
        hints.append(
            "Stagger overlapping leave -- several staff off the same day tightens "
            "the daily floors.")
        if S.alternating_weekends:
            hints.append(
                "Leave overlapping weekends constrains the weekend alternation -- "
                "turn the toggle off for this month or shift the leave.")
    return hints


# ---------------------------------------------------------------------------
# MODEL + SOLVE
# ---------------------------------------------------------------------------

@dataclass
class Solution:
    grid: list[list[str]]      # grid[employee][day] in {D, N, O, V}
    status: str


def build_and_solve(settings: ScheduleSettings) -> Solution:
    S = settings
    days = S.days
    n = len(S.employees)
    elig = night_eligible_indices(S)            # may work nights
    elig_set = set(elig)
    day_cap = day_capable_indices(S)            # may work day shifts
    day_cap_set = set(day_cap)
    weekend_days = weekend_day_indices(S)
    weekend_pairs = weekend_pair_indices(S)
    lv = leave_day_sets(S)                      # per-employee leave day indices
    tgt = shift_targets(S)                      # exact totals, prorated by leave
    segs = available_segments(S)                # per-employee free stretches
    has_leave = any(lv)                         # gates every leave-only branch

    model = cp_model.CpModel()

    work = {(e, d): model.NewBoolVar(f"work_{e}_{d}")
            for e in range(n) for d in range(days)}
    night = {(e, d): model.NewBoolVar(f"night_{e}_{d}")
             for e in range(n) for d in range(days)}

    for e in range(n):
        for d in range(days):
            model.Add(night[(e, d)] <= work[(e, d)])

    # Only night-eligible staff may work nights (rotate mode => everyone).
    for e in range(n):
        if e not in elig_set:
            for d in range(days):
                model.Add(night[(e, d)] == 0)

    # Leave days are hard-blocked: no work may be scheduled on them (night <=
    # work zeroes the night for free). Adds no constraint when leave is empty.
    for e in range(n):
        for d in sorted(lv[e]):
            model.Add(work[(e, d)] == 0)

    # Fixed nights-only team: every working day is a night.
    if not S.is_rotate and S.night_team_nights_only:
        for e in elig_set:
            for d in range(days):
                model.Add(work[(e, d)] == night[(e, d)])

    # EXACT monthly workload per employee (the target is shifts_per_employee
    # prorated by that employee's leave days; identical when no leave).
    for e in range(n):
        model.Add(sum(work[(e, d)] for d in range(days)) == tgt[e])

    # Daily staffing bands. Night coverage is a band [night_min, night_max]:
    # overlap above the floor appears only where the exact totals force it (in
    # fixed_team mode), never below night_min (a night is never uncovered).
    night_count = {}
    for d in range(days):
        day_count = sum(work[(e, d)] - night[(e, d)] for e in range(n))
        model.Add(day_count >= S.day_min)
        model.Add(day_count <= S.day_max)
        ncv = model.NewIntVar(0, n, f"nightcount_{d}")
        model.Add(ncv == sum(night[(e, d)] for e in range(n)))
        model.Add(ncv >= S.night_min)
        model.Add(ncv <= S.night_max)
        night_count[d] = ncv

    # No Night -> Day the next calendar day (post-night recovery).
    for e in range(n):
        for d in range(days - 1):
            day_shift_next = work[(e, d + 1)] - night[(e, d + 1)]
            model.Add(night[(e, d)] + day_shift_next <= 1)

    # Run-length rules, applied WITHIN each free stretch (leave days belong to
    # no run; with no leave the single stretch (0, days-1) reproduces the exact
    # pre-leave constraints). The min-block encoding is specific to a minimum
    # of 2. The max-work/max-night windows stay global: a window crossing leave
    # contains a forced 0, so global windows == per-segment semantics there.
    assert S.min_consec_work == 2 and S.min_consec_off == 2, (
        "build_and_solve() only encodes a minimum block length of 2.")
    for e in range(n):
        if has_leave and tgt[e] == 0:
            # Leave >= target: the exact total already forces every day off, and
            # the off-run rules cannot apply to someone with no workable pattern.
            continue
        w = [work[(e, d)] for d in range(days)]

        # Max consecutive WORKING days (combined day+night run cap).
        for start in range(days - S.max_consec_work):
            model.Add(sum(w[start:start + S.max_consec_work + 1]) <= S.max_consec_work)

        # Max consecutive NIGHTS (priority cap; binds night workers to <= 3).
        for start in range(days - S.max_consec_night):
            model.Add(sum(night[(e, d)] for d in range(start, start + S.max_consec_night + 1))
                      <= S.max_consec_night)

        # Max consecutive OFF days: any window of (MAX_OFF+1) days INSIDE one
        # stretch has >= 1 work (a window may never span leave -- the leave
        # itself is the long absence, not a rostered off run).
        for (a, b) in segs[e]:
            for start in range(a, b - S.max_consec_off + 1):
                model.Add(sum(w[start:start + S.max_consec_off + 1]) >= 1)

        # Min consecutive working days = 2 (no isolated single work day), the
        # stretch edges playing the month-edge role. A 1-day stretch cannot
        # host a legal 2-run, so it is forced off (validator exempts it).
        for (a, b) in segs[e]:
            if a == b:
                model.Add(w[a] == 0)
                continue
            model.Add(w[a] <= w[a + 1])
            model.Add(w[b] <= w[b - 1])
            for d in range(a + 1, b):
                model.Add(w[d] <= w[d - 1] + w[d + 1])

        # Min consecutive off days = 2 (no isolated single off day).
        for (a, b) in segs[e]:
            if a == b:
                continue
            model.Add(w[a + 1] <= w[a])
            model.Add(w[b - 1] <= w[b])
            for d in range(a + 1, b):
                model.Add(w[d - 1] + w[d + 1] - w[d] <= 1)

    # ---- Soft fairness objectives (four equally-weighted goals) ----------
    # Each goal minimises a spread (max - min) so the burden lands evenly. They
    # are SOFT: they shape the chosen roster but never make a feasible rule set
    # infeasible. The validator re-derives each from the finished grid.

    def minmax_gap(values, name, ub):
        hi = model.NewIntVar(0, ub, name + "_hi")
        lo = model.NewIntVar(0, ub, name + "_lo")
        for c in values:
            model.Add(hi >= c)
            model.Add(lo <= c)
        return hi - lo

    # Signed variant for the leave-only weighted deviations: those go negative,
    # and minmax_gap's [0, ub] bounds would make any negative value infeasible
    # -- never reuse minmax_gap for them.
    def signed_gap(values, name, ub):
        hi = model.NewIntVar(-ub, ub, name + "_hi")
        lo = model.NewIntVar(-ub, ub, name + "_lo")
        for c in values:
            model.Add(hi >= c)
            model.Add(lo <= c)
        return hi - lo

    def weighted_devs(pool, exprs, per_cap):
        """Division-free target-weighted deviations over one pool: with T = the
        pool's summed targets and X = the pool's summed values, dev_e =
        T*x_e - tgt[e]*X is 0 exactly when everyone sits at their target share
        (so a leave-taker is not a permanent outlier). Leave months only --
        validate() re-derives the identical quantity. Returns (devs, bound)."""
        T = sum(tgt[p] for p in pool)
        X = sum(exprs) if exprs else 0
        devs = [T * x - tgt[p] * X for p, x in zip(pool, exprs)]
        return devs, (T + n * days) * max(1, per_cap)

    T_all = sum(tgt)

    # Goal 1: equal total shifts (already pinned exact; kept for robustness).
    # With leave, measure each total against that employee's own target.
    totals = [sum(work[(e, d)] for d in range(days)) for e in range(n)]
    if has_leave:
        gap_total = minmax_gap([totals[e] - tgt[e] for e in range(n)], "tot", days)
    else:
        gap_total = minmax_gap(totals, "tot", days)

    # Goal 2: balanced night load across night-eligible staff.
    night_exprs = [sum(night[(e, d)] for d in range(days)) for e in elig]
    if has_leave:
        devs, ub = weighted_devs(elig, night_exprs, days)
        gap_night = signed_gap(devs or [0], "night", ub)
    else:
        gap_night = minmax_gap(night_exprs or [0], "night", days)

    # Goal 3: balanced weekends -- even Fri/Sat duty within each pool, plus a
    # full Fri+Sat weekend off for everyone (best effort).
    wknd = max(1, len(weekend_days))
    wknd_day = [sum(work[(e, d)] - night[(e, d)] for d in weekend_days) for e in day_cap]
    wknd_night = [sum(night[(e, d)] for d in weekend_days) for e in elig]
    if has_leave:
        devs_d, ub_d = weighted_devs(day_cap, wknd_day, wknd)
        devs_n, ub_n = weighted_devs(elig, wknd_night, wknd)
        gap_wknd = (signed_gap(devs_d or [0], "wkday", ub_d)
                    + signed_gap(devs_n or [0], "wknight", ub_n))
    else:
        gap_wknd = (minmax_gap(wknd_day or [0], "wkday", wknd)
                    + minmax_gap(wknd_night or [0], "wknight", wknd))
    has_full_off = []
    for e in range(n):
        bits = []
        for (f, s) in weekend_pairs:
            b = model.NewBoolVar(f"offwknd_{e}_{f}")
            model.Add(work[(e, f)] + work[(e, s)] == 0).OnlyEnforceIf(b)
            model.Add(work[(e, f)] + work[(e, s)] >= 1).OnlyEnforceIf(b.Not())
            bits.append(b)
        h = model.NewBoolVar(f"hasoffwknd_{e}")
        if bits:
            model.AddMaxEquality(h, bits)
        else:
            model.Add(h == 0)
        has_full_off.append(h)

        # HARD toggle -- strict alternating weekends. Reusing the same reified
        # `bits` (bit == 1 iff that full weekend is fully off), force off/on to
        # alternate: of every two consecutive full weekends exactly one is off.
        # Same `bits`, same weekend_pairs, same gating as validate(), so the
        # constraint and the independent check score ONE quantity. When the
        # toggle is False this adds no variable and no constraint, so the model
        # stays byte-identical to the pre-feature engine. Leave exemptions
        # (mirrored in validate()): a target-0 employee is exempt (leave pins
        # every bit to 1, so any kept pair would be unsatisfiable), and any pair
        # touching a weekend fully covered by that employee's leave is skipped.
        if S.alternating_weekends and not (has_leave and tgt[e] == 0):
            full_leave = [f in lv[e] and s in lv[e] for (f, s) in weekend_pairs]
            for w in range(len(bits) - 1):
                if full_leave[w] or full_leave[w + 1]:
                    continue
                model.Add(bits[w] + bits[w + 1] == 1)
    num_no_full_off = n - sum(has_full_off)
    # In leave months the duty spreads above are in pool-target units; scale the
    # people-count term to match so neither half of the goal drowns the other.
    weekend_term = gap_wknd + (T_all * num_no_full_off if has_leave
                               else num_no_full_off)

    # Goal 4: balanced "undesirable runs", spread evenly *within* each pool (a
    # fixed night team structurally carries all overlap, so a cross-pool
    # comparison is not a fairness signal -- mirror the weekend goal). Rough load
    # per person = overlap nights worked + count of max-length day runs + count
    # of max-length night runs. This is the EXACT quantity validate() re-derives
    # from the grid, so optimising it actually moves the validated metric (a run
    # that hits the cap can only be length == cap, so one window == one run).
    overlap = {}
    for d in range(days):
        b = model.NewBoolVar(f"overlap_{d}")          # a stacked night above floor
        model.Add(night_count[d] >= S.night_min + 1).OnlyEnforceIf(b)
        model.Add(night_count[d] <= S.night_min).OnlyEnforceIf(b.Not())
        overlap[d] = b

    Lw, Ln = S.max_consec_work, S.max_consec_night
    rough_terms = [0] * n
    for e in range(n):
        parts = []
        if e in elig_set:                              # overlap nights worked
            for d in range(days):
                ov = model.NewBoolVar(f"ovn_{e}_{d}")
                model.Add(ov <= night[(e, d)])
                model.Add(ov <= overlap[d])
                model.Add(ov >= night[(e, d)] + overlap[d] - 1)
                parts.append(ov)
        if e in day_cap_set:                           # max-length day runs
            for start in range(days - Lw + 1):
                seg = [work[(e, d)] - night[(e, d)] for d in range(start, start + Lw)]
                wmax = model.NewBoolVar(f"wmax_{e}_{start}")
                model.Add(sum(seg) >= Lw).OnlyEnforceIf(wmax)
                model.Add(sum(seg) <= Lw - 1).OnlyEnforceIf(wmax.Not())
                parts.append(wmax)
        if e in elig_set:                              # max-length night runs
            for start in range(days - Ln + 1):
                seg = [night[(e, d)] for d in range(start, start + Ln)]
                nmax = model.NewBoolVar(f"nmax_{e}_{start}")
                model.Add(sum(seg) >= Ln).OnlyEnforceIf(nmax)
                model.Add(sum(seg) <= Ln - 1).OnlyEnforceIf(nmax.Not())
                parts.append(nmax)
        rough_terms[e] = sum(parts) if parts else 0
    # Sum the per-pool spreads -- an upper bound on the max-pool spread the
    # validator checks, so minimising it drives the validated metric down. In
    # leave months each pool's raw spread is priced in that pool's summed-target
    # units (the same scale as the weighted goals) so the four goals stay
    # commensurate; positive scaling is monotone, so the validator's raw check
    # still scores the same quantity.
    pools = [elig] + ([day_cap] if set(day_cap) != elig_set else [])
    if has_leave:
        gap_runs = sum(sum(tgt[p] for p in pool)
                       * minmax_gap([rough_terms[e] for e in pool] or [0],
                                    f"rough{k}", 2 * days)
                       for k, pool in enumerate(pools))
    else:
        gap_runs = sum(minmax_gap([rough_terms[e] for e in pool] or [0],
                                  f"rough{k}", 2 * days)
                       for k, pool in enumerate(pools))

    model.Minimize(
        S.w_fair_total * (T_all if has_leave else 1) * gap_total
        + S.w_fair_night * gap_night
        + S.w_fair_weekend * weekend_term
        + S.w_fair_runs * gap_runs
    )

    solver = cp_model.CpSolver()
    solver.parameters.max_deterministic_time = S.solver_det_time_limit
    solver.parameters.max_time_in_seconds = S.solver_time_limit
    solver.parameters.num_search_workers = 1
    solver.parameters.random_seed = 42

    status = solver.Solve(model)
    status_name = solver.StatusName(status)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return Solution(grid=[], status=status_name)

    grid: list[list[str]] = []
    for e in range(n):
        row = []
        for d in range(days):
            if solver.Value(work[(e, d)]) == 0:
                # Label configured leave only when the hard block actually held:
                # a worked leave day would surface as D/N below and loudly fail
                # the "Leave days are honored" check (constraint and check score
                # one quantity -- never label from settings alone).
                row.append(LEAVE if d in lv[e] else OFF)
            elif solver.Value(night[(e, d)]) == 1:
                row.append(NIGHT)
            else:
                row.append(DAY)
        grid.append(row)
    return Solution(grid=grid, status=status_name)


# ---------------------------------------------------------------------------
# INDEPENDENT VALIDATOR
# ---------------------------------------------------------------------------

@dataclass
class RuleResult:
    name: str
    passed: bool
    measured: str
    # --- Optional structured display metadata (set for fairness goals) ------
    # Purely additive: lets a front-end render the spread/tolerance beside each
    # fairness goal and split hard vs fairness checks without string-parsing
    # `name`/`measured`. Never consulted by the solver or the validator's
    # pass/fail logic -- the booleans above remain the single source of truth.
    kind: str = "hard"                 # "hard" | "fairness"
    spread: int | None = None          # measured spread (max - min) for fairness
    tolerance: int | None = None       # tolerance the spread is checked against


def _runs(seq: list[bool]) -> list[int]:
    lengths, current = [], 0
    for v in seq:
        if v:
            current += 1
        elif current:
            lengths.append(current)
            current = 0
    if current:
        lengths.append(current)
    return lengths


def employee_loads(settings: ScheduleSettings, grid: list[list[str]]) -> list[dict]:
    """Per-employee fairness facts, re-derived from the grid (used by the UI,
    the Excel summary, and the fairness validator so they all agree)."""
    S = settings
    days = S.days
    weekend_days = weekend_day_indices(S)
    weekend_pairs = weekend_pair_indices(S)
    lv = leave_day_sets(S)
    tgt = shift_targets(S)
    night_counts = [sum(1 for e in range(len(S.employees)) if grid[e][d] == NIGHT)
                    for d in range(days)]
    loads = []
    for e in range(len(S.employees)):
        row = grid[e]
        day_runs = _runs([row[d] == DAY for d in range(days)])
        night_runs = _runs([row[d] == NIGHT for d in range(days)])
        overlaps = sum(1 for d in range(days)
                       if row[d] == NIGHT and night_counts[d] >= S.night_min + 1)
        max_runs = (sum(1 for r in day_runs if r >= S.max_consec_work)
                    + sum(1 for r in night_runs if r >= S.max_consec_night))
        loads.append({
            "name": S.employees[e],
            "night_member": is_night_member(S, e),
            "total": sum(1 for d in range(days) if row[d] not in (OFF, LEAVE)),
            "day": sum(1 for d in range(days) if row[d] == DAY),
            "night": sum(1 for d in range(days) if row[d] == NIGHT),
            "weekend": sum(1 for d in weekend_days if row[d] not in (OFF, LEAVE)),
            "overlaps": overlaps,
            "max_runs": max_runs,
            "rough": overlaps + max_runs,
            # A fully-leave weekend counts as that person's weekend off (they
            # ARE off) -- matches the solver's reified weekend bit exactly.
            "weekends_off": sum(1 for (f, s) in weekend_pairs
                                if row[f] in (OFF, LEAVE) and row[s] in (OFF, LEAVE)),
            "leave": len(lv[e]),
            "target": tgt[e],
        })
    return loads


def coverage_per_day(settings: ScheduleSettings, grid: list[list[str]]) -> list[dict]:
    """Per-day staffing counts, re-derived from the grid (display only).

    Independent re-derivation -- reads only the finished grid (like
    employee_loads), never the solver or the validator. Lets a front-end draw a
    per-day coverage strip without recomputing any rule. `over_floor` flags a
    day carrying a stacked night above night_min (a forced-overlap day).
    """
    S = settings
    n = len(S.employees)
    out = []
    for d in range(S.days):
        day_n = sum(1 for e in range(n) if grid[e][d] == DAY)
        night_n = sum(1 for e in range(n) if grid[e][d] == NIGHT)
        leave_n = sum(1 for e in range(n) if grid[e][d] == LEAVE)
        out.append({
            "day_index": d,
            "label": f"D{d + 1}",
            "weekday": weekday_of(S, d),
            "is_weekend": weekday_of(S, d) in S.weekend_days,
            "day": day_n,
            "night": night_n,
            "leave": leave_n,
            "off": n - day_n - night_n - leave_n,
            "over_floor": night_n >= S.night_min + 1,
        })
    return out


def validate(settings: ScheduleSettings, grid: list[list[str]]) -> list[RuleResult]:
    S = settings
    days = S.days
    n = len(S.employees)
    rotate = S.is_rotate
    elig = night_eligible_indices(S)
    elig_set = set(elig)
    day_cap = day_capable_indices(S)
    weekend_days = weekend_day_indices(S)
    weekend_pairs = weekend_pair_indices(S)
    lv = leave_day_sets(S)                      # re-derived from settings, never
    tgt = shift_targets(S)                      # from the solver
    segs = available_segments(S)
    has_leave = any(lv)
    results: list[RuleResult] = []

    day_counts = [sum(1 for e in range(n) if grid[e][d] == DAY) for d in range(days)]
    night_counts = [sum(1 for e in range(n) if grid[e][d] == NIGHT) for d in range(days)]

    # ===================== HARD RULES =====================================
    results.append(RuleResult(
        f"Day staffing per day within [{S.day_min}, {S.day_max}]",
        all(S.day_min <= c <= S.day_max for c in day_counts),
        f"min={min(day_counts)}, max={max(day_counts)}"))
    results.append(RuleResult(
        f"Night coverage every day >= {S.night_min} (<= {S.night_max})",
        all(S.night_min <= c <= S.night_max for c in night_counts),
        f"min={min(night_counts)}, max={max(night_counts)}"))

    # Forced-overlap minimality only has a fixed target with a nights-only team.
    if not rotate and S.night_team_nights_only:
        expected = expected_double_nights(S)
        actual = sum(max(0, c - S.night_min) for c in night_counts)
        results.append(RuleResult(
            "Night overlap is minimal (only forced extra nights)",
            actual == expected,
            f"overlap-nights above floor={actual} (forced minimum={expected})"))

    results.append(RuleResult(
        "At most one shift per employee per day",
        all(grid[e][d] in (DAY, NIGHT, OFF, LEAVE) for e in range(n) for d in range(days)),
        "one code per cell"))

    # Leave days are honored: V appears exactly on the configured leave days,
    # both directions. Extraction labels LEAVE only where the solver's work
    # value is 0, so a violated hard-block surfaces as D/N and fails HERE.
    if has_leave:
        mismatches = [(S.employees[e], f"D{d + 1}") for e in range(n)
                      for d in range(days) if (grid[e][d] == LEAVE) != (d in lv[e])]
        results.append(RuleResult(
            "Leave days are honored",
            not mismatches,
            f"{sum(len(x) for x in lv)} leave day(s) placed exactly as configured"
            if not mismatches else f"{len(mismatches)} mismatch(es): {mismatches[:6]}"))

    totals = [sum(1 for d in range(days) if grid[e][d] not in (OFF, LEAVE))
              for e in range(n)]
    if has_leave:
        results.append(RuleResult(
            "Shift totals match each employee's prorated target",
            all(totals[e] == tgt[e] for e in range(n)),
            f"totals={totals} targets={tgt}"))
    else:
        results.append(RuleResult(
            f"Exactly {S.shifts_per_employee} shifts per employee",
            all(t == S.shifts_per_employee for t in totals),
            f"totals={totals}"))

    night_workers = {S.employees[e] for e in range(n)
                     if any(grid[e][d] == NIGHT for d in range(days))}
    if rotate:
        # No dedicated team: nights may land on anyone -- just confirm that.
        results.append(RuleResult(
            "All staff are night-eligible (rotation)",
            night_workers.issubset(set(S.employees)),
            f"{len(night_workers)} of {n} staff worked >=1 night"))
    else:
        results.append(RuleResult(
            "Only the night team works nights",
            night_workers.issubset(set(S.night_team)),
            f"night workers={sorted(night_workers)}"))
        if S.night_team_nights_only:
            # With a nights-only team, every member works only nights, so all of
            # them necessarily carry >=1 night. (If the team could also work days
            # a member legitimately taking zero nights must NOT count as a fault,
            # so this check only applies in the nights-only case.) A target-0
            # member (leave >= target) legitimately works zero nights.
            n_carriers = (sum(1 for e in elig if tgt[e] > 0) if has_leave
                          else len(S.night_team))
            results.append(RuleResult(
                f"Exactly {n_carriers} people carry nights for the month",
                len(night_workers) == n_carriers,
                f"{len(night_workers)} have >=1 night: {sorted(night_workers)}"))
            nightteam_dayshifts = [S.employees[e] for e in elig_set
                                   if any(grid[e][d] == DAY for d in range(days))]
            results.append(RuleResult(
                "Night team works nights only (no day shifts)",
                not nightteam_dayshifts,
                "0 day shifts by night team" if not nightteam_dayshifts
                else f"violations: {nightteam_dayshifts}"))

    nd = [(S.employees[e], f"D{d + 1}->D{d + 2}") for e in range(n)
          for d in range(days - 1) if grid[e][d] == NIGHT and grid[e][d + 1] == DAY]
    results.append(RuleResult(
        "No Night->Day next-day transition",
        not nd,
        f"{len(nd)} transitions" + ("" if not nd else f": {nd[:6]}")))

    # The solver caps COMBINED working runs (day or night, since a run can mix
    # both shapes outside fixed_team's nights-only team -- see build_and_solve's
    # max_consec_work window over `work`, not just DAY cells). Re-derive the same
    # quantity here with the not-in-(OFF, LEAVE) predicate the min-work check
    # below already uses: a window of max_consec_work+1 cells with no O/V cell IS
    # a working run longer than the cap, so longest-run and the solver's window
    # encoding agree exactly.
    max_work_run = max((max(_runs([grid[e][d] not in (OFF, LEAVE) for d in range(days)]),
                           default=0)
                        for e in range(n)), default=0)
    results.append(RuleResult(
        f"Max {S.max_consec_work} consecutive working days",
        max_work_run <= S.max_consec_work,
        f"longest working streak={max_work_run}"))

    max_night_run = max((max(_runs([grid[e][d] == NIGHT for d in range(days)]), default=0)
                         for e in range(n)), default=0)
    results.append(RuleResult(
        f"Max {S.max_consec_night} consecutive nights",
        max_night_run <= S.max_consec_night,
        f"longest night streak={max_night_run}"))

    # The off/work run rules live WITHIN each free stretch: a LEAVE cell is
    # neither work nor off, so runs computed on exact codes already split at
    # leave. Target-0 employees (leave >= target) are exempt -- their forced
    # all-off month is not a rostered pattern.
    exempt = {e for e in range(n) if has_leave and tgt[e] == 0}

    min_work_run = min((min(_runs([grid[e][d] not in (OFF, LEAVE) for d in range(days)]),
                            default=S.min_consec_work)
                        for e in range(n) if e not in exempt), default=S.min_consec_work)
    results.append(RuleResult(
        f"Min {S.min_consec_work} consecutive working days",
        min_work_run >= S.min_consec_work,
        f"shortest work block={min_work_run}"))

    # Min-off exemption: an off-run exactly filling a whole stretch shorter than
    # the minimum (a singleton day squeezed between leave/month edges) is the
    # solver's only legal output there -- forced off, not an isolated off day.
    def _min_off(e):
        runs = []
        for (a, b) in segs[e]:
            cells = [grid[e][d] == OFF for d in range(a, b + 1)]
            if (b - a + 1) < S.min_consec_off and all(cells):
                continue
            runs.extend(_runs(cells))
        return min(runs, default=S.min_consec_off)

    min_off_run = min((_min_off(e) for e in range(n) if e not in exempt),
                      default=S.min_consec_off)
    results.append(RuleResult(
        f"Min {S.min_consec_off} consecutive off days",
        min_off_run >= S.min_consec_off,
        f"shortest off block={min_off_run}"))

    max_off_run = max((max(_runs([grid[e][d] == OFF for d in range(days)]), default=0)
                       for e in range(n) if e not in exempt), default=0)
    results.append(RuleResult(
        f"Max {S.max_consec_off} consecutive off days",
        max_off_run <= S.max_consec_off,
        f"longest off streak={max_off_run}"))

    # Alternating weekends (HARD, present only when the toggle is on). Re-derived
    # from the grid alone: a full weekend is "off" iff the employee is OFF on both
    # its Fri and Sat; strict alternation is violated whenever two CONSECUTIVE
    # full weekends share the same on/off state. Mirrors the model constraint
    # exactly (same weekend_pairs, same off-both-days test, same gating), so a
    # solver OPTIMAL and this check can never disagree. Vacuous (<2 full weekends).
    if S.alternating_weekends:
        # Leave exemptions mirror the constraint exactly: target-0 employees are
        # exempt, pairs touching a fully-leave weekend are skipped, and a
        # weekend is "off" when both days are OFF or LEAVE (the solver bit).
        alt_viol = []
        for e in range(n):
            if e in exempt:
                continue
            off = [grid[e][f] in (OFF, LEAVE) and grid[e][s] in (OFF, LEAVE)
                   for (f, s) in weekend_pairs]
            full_leave = [f in lv[e] and s in lv[e] for (f, s) in weekend_pairs]
            for w in range(len(off) - 1):
                if full_leave[w] or full_leave[w + 1]:
                    continue
                if off[w] == off[w + 1]:
                    f1, f2 = weekend_pairs[w][0], weekend_pairs[w + 1][0]
                    alt_viol.append((S.employees[e], f"D{f1 + 1}&D{f2 + 1}"))
        results.append(RuleResult(
            "Weekends alternate off / on for every employee",
            not alt_viol,
            f"{len(weekend_pairs)} full weekends; alternation holds" if not alt_viol
            else f"{len(alt_viol)} violation(s): {alt_viol[:6]}"))

    # ===================== FAIRNESS (4 soft goals) ========================
    # Each reports PASS/FAIL against a tunable tolerance, plus the spread.
    loads = employee_loads(S, grid)

    def spread(values):
        vals = list(values) or [0]
        return max(vals) - min(vals)

    # F1 -- equal total shifts (with leave: each vs their own prorated target).
    if has_leave:
        sp_total = spread(loads[e]["total"] - tgt[e] for e in range(n))
        results.append(RuleResult(
            f"Fairness - equal total shifts (spread <= {S.fair_tol_total})",
            sp_total <= S.fair_tol_total,
            f"spread={sp_total} (every total matches its prorated target)",
            kind="fairness", spread=sp_total, tolerance=S.fair_tol_total))
    else:
        sp_total = spread(ld["total"] for ld in loads)
        results.append(RuleResult(
            f"Fairness - equal total shifts (spread <= {S.fair_tol_total})",
            sp_total <= S.fair_tol_total,
            f"spread={sp_total} (every total = {S.shifts_per_employee})",
            kind="fairness", spread=sp_total, tolerance=S.fair_tol_total))

    # F2 -- balanced night load across night-eligible staff. With leave the
    # spread is over the target-weighted deviations T*x_e - tgt[e]*X -- the
    # EXACT quantity the solver's goal-2 term minimises -- against tol * T.
    night_by = {loads[e]["name"]: loads[e]["night"] for e in elig}
    if has_leave:
        t_night = sum(tgt[e] for e in elig)
        x_night = sum(loads[e]["night"] for e in elig)
        sp_night = spread(t_night * loads[e]["night"] - tgt[e] * x_night
                          for e in elig)
        tol_night = S.fair_tol_night * t_night
        results.append(RuleResult(
            f"Fairness - balanced night load (weighted spread <= {tol_night})",
            sp_night <= tol_night,
            f"weighted spread={sp_night} (pool target={t_night})  "
            f"nights/eligible={night_by}",
            kind="fairness", spread=sp_night, tolerance=tol_night))
    else:
        sp_night = spread(night_by.values())
        results.append(RuleResult(
            f"Fairness - balanced night load (spread <= {S.fair_tol_night})",
            sp_night <= S.fair_tol_night,
            f"spread={sp_night}  nights/eligible={night_by}",
            kind="fairness", spread=sp_night, tolerance=S.fair_tol_night))

    # F3 -- balanced weekends: even Fri/Sat duty in each pool + a weekend off each.
    no_weekend_off = [loads[e]["name"] for e in range(n)
                      if weekend_pairs and loads[e]["weekends_off"] == 0]
    if has_leave:
        # Per-pool weighted tests (the pools have different target sums, so one
        # shared bound would be wrong for whichever pool's T wasn't used). The
        # reported spread/tolerance pair comes from the binding pool so the two
        # numbers stay in one unit.
        t_day = sum(tgt[e] for e in day_cap)
        t_ni = sum(tgt[e] for e in elig)
        xd = {e: sum(1 for d in weekend_days if grid[e][d] == DAY) for e in day_cap}
        xn = {e: sum(1 for d in weekend_days if grid[e][d] == NIGHT) for e in elig}
        x_day, x_ni = sum(xd.values()), sum(xn.values())
        sp_wknd_day = spread(t_day * xd[e] - tgt[e] * x_day for e in day_cap)
        sp_wknd_night = spread(t_ni * xn[e] - tgt[e] * x_ni for e in elig)
        day_ok = sp_wknd_day <= S.fair_tol_weekend * t_day
        night_ok = sp_wknd_night <= S.fair_tol_weekend * t_ni
        passed_wknd = day_ok and night_ok and not no_weekend_off
        if day_ok != night_ok:
            bind_sp, bind_t = ((sp_wknd_night, t_ni) if day_ok
                               else (sp_wknd_day, t_day))
        else:
            bind_sp, bind_t = ((sp_wknd_day, t_day)
                               if sp_wknd_day * max(1, t_ni) >= sp_wknd_night * max(1, t_day)
                               else (sp_wknd_night, t_ni))
        results.append(RuleResult(
            "Fairness - balanced weekends (weighted spread per pool, a weekend off each)",
            passed_wknd,
            f"day dev spread={sp_wknd_day} (tol {S.fair_tol_weekend * t_day}), "
            f"night dev spread={sp_wknd_night} (tol {S.fair_tol_weekend * t_ni}), "
            f"no weekend-off for: {no_weekend_off or 'none'}",
            kind="fairness", spread=bind_sp,
            tolerance=S.fair_tol_weekend * bind_t))
    else:
        sp_wknd_day = spread(sum(1 for d in weekend_days if grid[e][d] == DAY) for e in day_cap)
        sp_wknd_night = spread(sum(1 for d in weekend_days if grid[e][d] == NIGHT) for e in elig)
        passed_wknd = (max(sp_wknd_day, sp_wknd_night) <= S.fair_tol_weekend
                       and not no_weekend_off)
        results.append(RuleResult(
            f"Fairness - balanced weekends (spread <= {S.fair_tol_weekend}, a weekend off each)",
            passed_wknd,
            f"day spread={sp_wknd_day}, night spread={sp_wknd_night}, "
            f"no weekend-off for: {no_weekend_off or 'none'}",
            kind="fairness", spread=max(sp_wknd_day, sp_wknd_night),
            tolerance=S.fair_tol_weekend))

    # F4 -- balanced undesirable runs *within* each pool (a fixed night team
    # structurally carries all overlap; comparing it to the day team is not a
    # fairness signal -- balance is what matters inside each group).
    rough_by = {ld["name"]: ld["rough"] for ld in loads}
    pools = [elig] + ([day_cap] if set(day_cap) != elig_set else [])
    sp_rough = max((spread(loads[e]["rough"] for e in pool) for pool in pools), default=0)
    results.append(RuleResult(
        f"Fairness - balanced undesirable runs (within-pool spread <= {S.fair_tol_runs})",
        sp_rough <= S.fair_tol_runs,
        f"within-pool spread={sp_rough}  rough-load(overlap+max-runs)={rough_by}",
        kind="fairness", spread=sp_rough, tolerance=S.fair_tol_runs))

    return results


# ---------------------------------------------------------------------------
# EXCEL EXPORT
# ---------------------------------------------------------------------------

FILL_DAY = PatternFill("solid", fgColor="C6EFCE")
FILL_NIGHT = PatternFill("solid", fgColor="BDD7EE")
FILL_LEAVE = PatternFill("solid", fgColor="FFC7CE")  # red now marks leave, not off
FILL_PASS = PatternFill("solid", fgColor="C6EFCE")
FILL_FAIL = PatternFill("solid", fgColor="FFC7CE")
FILL_HEADER = PatternFill("solid", fgColor="305496")
FILL_WEEKEND = PatternFill("solid", fgColor="FFF2CC")
FILL_AM = PatternFill("solid", fgColor="FFE699")    # AM shift (fixed-pattern staff)

FONT_HEADER = Font(bold=True, color="FFFFFF")
FONT_BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# OFF cells are truly blank (no text, no fill -- _style skips a None fill);
# leave cells carry the red that used to mean off.
CELL_LABEL = {DAY: "D", NIGHT: "N", OFF: "", LEAVE: "V", AM: "AM"}
CELL_FILL = {DAY: FILL_DAY, NIGHT: FILL_NIGHT, OFF: None, LEAVE: FILL_LEAVE, AM: FILL_AM}

# Hex colors reused by the Streamlit grid so the two views match.
WEB_COLORS = {DAY: "#C6EFCE", NIGHT: "#BDD7EE", OFF: "#FFFFFF",
              LEAVE: "#FFC7CE", AM: "#FFE699"}


def _style(cell, value="", fill=None, font=None, align=CENTER, border=True):
    cell.value = value
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.alignment = align
    if border:
        cell.border = BORDER


def build_workbook(settings: ScheduleSettings, grid: list[list[str]],
                   results: list[RuleResult]) -> Workbook:
    wb = Workbook()
    _write_roster(settings, wb.active, grid)
    _write_summary(settings, wb.create_sheet("Summary"), grid, results)
    return wb


def export(settings: ScheduleSettings, grid: list[list[str]],
           results: list[RuleResult], path: str = "schedule.xlsx") -> None:
    build_workbook(settings, grid, results).save(path)


def export_bytes(settings: ScheduleSettings, grid: list[list[str]],
                 results: list[RuleResult]) -> bytes:
    buf = BytesIO()
    build_workbook(settings, grid, results).save(buf)
    return buf.getvalue()


def _write_roster(settings: ScheduleSettings, ws, grid: list[list[str]]) -> None:
    S = settings
    days = S.days
    ws.title = "Schedule"
    weekend = set(weekend_day_indices(S))

    c = ws.cell(1, 1)
    _style(c, f"Employee\n{S.month_label}", FILL_HEADER, FONT_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for d in range(days):
        cell = ws.cell(1, 2 + d)
        is_wknd = d in weekend
        _style(cell, f"D{d + 1}\n{weekday_of(S, d)}",
               FILL_WEEKEND if is_wknd else FILL_HEADER,
               FONT_BOLD if is_wknd else FONT_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for j, label in enumerate(["Total", "Day", "Night", "Fri/Sat", "Hours"]):
        _style(ws.cell(1, 2 + days + j), label, FILL_HEADER, FONT_HEADER)

    # Stat columns are LIVE formulas over the day cells, so a hand edit in the
    # workbook recounts automatically. Fri/Sat columns are non-contiguous, so
    # that formula is an explicit per-weekend-cell sum. Leave ("V") never counts
    # as worked; Hours = Total * hours_per_shift.
    last_col = get_column_letter(1 + days)
    total_col = get_column_letter(2 + days)
    wknd_cols = [get_column_letter(2 + d) for d in weekend_day_indices(S)]

    def stat_formulas(r, codes):
        rng = f"B{r}:{last_col}{r}"
        total = "=" + "+".join(f'COUNTIF({rng},"{c}")' for c in codes)
        frisat = ("=" + "+".join(f'COUNTIF({col}{r},"{c}")'
                                 for col in wknd_cols for c in codes)
                  if wknd_cols else 0)
        return total, frisat

    for e in range(len(S.employees)):
        label = S.employees[e] + ("  (night team)" if is_night_member(S, e) else "")
        r = 2 + e
        _style(ws.cell(r, 1), label, font=FONT_BOLD,
               align=Alignment(horizontal="left", vertical="center"))
        for d in range(days):
            code = grid[e][d]
            _style(ws.cell(r, 2 + d), CELL_LABEL[code], CELL_FILL[code])
        total_f, frisat_f = stat_formulas(r, ("D", "N"))
        rng = f"B{r}:{last_col}{r}"
        stats = [total_f, f'=COUNTIF({rng},"D")', f'=COUNTIF({rng},"N")',
                 frisat_f, f"={total_col}{r}*{S.hours_per_shift}"]
        for j, val in enumerate(stats):
            _style(ws.cell(r, 2 + days + j), val, font=FONT_BOLD)

    # AM staff: appended after the scheduled roster as extra rows (fixed weekly
    # pattern, entirely outside the solver -- no coverage/fairness accounting).
    # Their Day/Night stat columns are 0 by definition; Total = AM days worked,
    # Fri/Sat = AM days landing on the weekend. Hours stays blank: the AM shift
    # length is not the solver shift length, and AM staff take no rostered leave.
    am = am_rows(S)
    base = 2 + len(S.employees)
    for a, arow in enumerate(am):
        r = base + a
        _style(ws.cell(r, 1), S.am_team[a] + "  (AM)", font=FONT_BOLD,
               align=Alignment(horizontal="left", vertical="center"))
        for d in range(days):
            code = arow[d]
            _style(ws.cell(r, 2 + d), CELL_LABEL[code], CELL_FILL[code])
        total_f, frisat_f = stat_formulas(r, ("AM",))
        for j, val in enumerate([total_f, 0, 0, frisat_f, ""]):
            _style(ws.cell(r, 2 + days + j), val, font=FONT_BOLD)

    legend_row = len(S.employees) + len(S.am_team) + 4
    _style(ws.cell(legend_row, 1), "Legend:", font=FONT_BOLD, border=False)
    legend_items = [(DAY, "Day shift"), (NIGHT, "Night shift"), (OFF, "Off")]
    if S.leave:
        legend_items.append((LEAVE, "Leave"))
    if S.am_team:
        legend_items.append((AM, "AM shift"))
    for j, (code, text) in enumerate(legend_items):
        _style(ws.cell(legend_row, 2 + j * 2), CELL_LABEL[code], CELL_FILL[code])
        _style(ws.cell(legend_row, 3 + j * 2), text, border=False,
               align=Alignment(horizontal="left", vertical="center"))
    _style(ws.cell(legend_row + 1, 1), roster_caption(S),
           font=FONT_BOLD, border=False, align=Alignment(horizontal="left", vertical="center"))
    _style(ws.cell(legend_row + 2, 1),
           "Total/Day/Night/Fri-Sat/Hours recount the day cells live, so hand "
           "edits update them; the Summary sheet and the app's validation "
           "reflect only the generated roster.",
           border=False, align=Alignment(horizontal="left", vertical="center"))

    # Staff on leave: one row per entered range, real date cells.
    if S.leave:
        lt_row = legend_row + 4
        _style(ws.cell(lt_row, 1), "Staff on leave", font=FONT_BOLD, border=False,
               align=Alignment(horizontal="left", vertical="center"))
        for j, h in enumerate(["Name", "From", "To"]):
            _style(ws.cell(lt_row + 1, 1 + j), h, FILL_HEADER, FONT_HEADER)
        for i, (name, d0, d1) in enumerate(S.leave):
            _style(ws.cell(lt_row + 2 + i, 1), name,
                   align=Alignment(horizontal="left", vertical="center"))
            for j, dt in enumerate([d0, d1]):
                cell = ws.cell(lt_row + 2 + i, 2 + j)
                _style(cell, dt)
                cell.number_format = "DD MMM YYYY"

    ws.column_dimensions["A"].width = 18
    for d in range(days):
        ws.column_dimensions[get_column_letter(2 + d)].width = 6
    for j in range(5):
        ws.column_dimensions[get_column_letter(2 + days + j)].width = 8
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "B2"


def _write_summary(settings: ScheduleSettings, ws, grid: list[list[str]],
                   results: list[RuleResult]) -> None:
    S = settings
    days = S.days
    elig = night_eligible_indices(S)
    loads = employee_loads(S, grid)
    row = 1

    def header(text):
        nonlocal row
        _style(ws.cell(row, 1), text, FILL_HEADER, FONT_HEADER,
               align=Alignment(horizontal="left", vertical="center"))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

    header(f"Per-employee fairness summary - {S.month_label}")
    for j, h in enumerate(["Employee", "Total", "Day", "Night", "Fri/Sat", "Rough"]):
        _style(ws.cell(row, 1 + j), h, FILL_HEADER, FONT_HEADER)
    row += 1
    for e, ld in enumerate(loads):
        name = ld["name"] + ("  (night team)" if ld["night_member"] else "")
        cells = [name, ld["total"], ld["day"], ld["night"], ld["weekend"], ld["rough"]]
        for j, val in enumerate(cells):
            _style(ws.cell(row, 1 + j), val,
                   align=Alignment(horizontal="left" if j == 0 else "center", vertical="center"))
        row += 1
    row += 1

    header("Night coverage")
    split = {loads[e]["name"]: loads[e]["night"] for e in elig}
    for name, nval in split.items():
        _style(ws.cell(row, 1), name, align=Alignment(horizontal="left", vertical="center"))
        _style(ws.cell(row, 2), f"{nval} nights")
        row += 1
    if split:
        _style(ws.cell(row, 1), "Gap", font=FONT_BOLD, align=Alignment(horizontal="left", vertical="center"))
        _style(ws.cell(row, 2), max(split.values()) - min(split.values()), font=FONT_BOLD)
        row += 1
    overlap_days = sum(1 for d in range(days)
                       if sum(1 for e in range(len(S.employees)) if grid[e][d] == NIGHT)
                       >= S.night_min + 1)
    _style(ws.cell(row, 1), "Overlap-night days",
           font=FONT_BOLD, align=Alignment(horizontal="left", vertical="center"))
    _style(ws.cell(row, 2), overlap_days, font=FONT_BOLD)
    row += 2

    header("Rule validation")
    for j, h in enumerate(["Rule", "Result", "Measured"]):
        _style(ws.cell(row, 1 + j), h, FILL_HEADER, FONT_HEADER)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    row += 1
    for r in results:
        _style(ws.cell(row, 1), r.name, align=Alignment(horizontal="left", vertical="center"))
        _style(ws.cell(row, 2), "PASS" if r.passed else "FAIL",
               FILL_PASS if r.passed else FILL_FAIL, FONT_BOLD)
        _style(ws.cell(row, 3), r.measured, align=Alignment(horizontal="left", vertical="center"))
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        row += 1

    ws.column_dimensions["A"].width = 52
    for col in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 14


# ---------------------------------------------------------------------------
# MAIN (CLI)
# ---------------------------------------------------------------------------

def print_summary(settings: ScheduleSettings, grid: list[list[str]],
                  results: list[RuleResult]) -> bool:
    S = settings
    loads = employee_loads(S, grid)
    has_leave = any(ld["leave"] for ld in loads)
    print(f"\n{roster_caption(S)}")
    print("\nPer-employee counts:")
    lv_head = f" {'Leave':>6} {'Target':>7}" if has_leave else ""
    print(f"  {'Employee':<12} {'Total':>5} {'Day':>4} {'Night':>6} {'Fri/Sat':>8} "
          f"{'Rough':>6}{lv_head}")
    for ld in loads:
        tag = "  <- night team" if ld["night_member"] else ""
        lv_cols = f" {ld['leave']:>6} {ld['target']:>7}" if has_leave else ""
        print(f"  {ld['name']:<12} {ld['total']:>5} {ld['day']:>4} {ld['night']:>6} "
              f"{ld['weekend']:>8} {ld['rough']:>6}{lv_cols}{tag}")

    print("\nRule validation:")
    all_pass = True
    for r in results:
        all_pass &= r.passed
        print(f"  [{'PASS' if r.passed else 'FAIL'}] {r.name:<58} | {r.measured}")
    return all_pass


def main() -> int:
    settings = ScheduleSettings()

    problems = preflight(settings)
    if problems:
        print("INFEASIBLE rule set - stopping before scheduling:")
        for p in problems:
            print(f"  - {p.message}\n    suggestion: {p.suggestion}")
        return 2

    solution = build_and_solve(settings)
    if not solution.grid:
        print(f"No schedule found (solver status: {solution.status}).")
        print("The constraints are jointly infeasible. Try:")
        for hint in relaxation_hints(settings):
            print(f"  - {hint}")
        return 2

    results = validate(settings, solution.grid)
    all_pass = print_summary(settings, solution.grid, results)
    export(settings, solution.grid, results)
    print(f"\nWrote schedule.xlsx (solver status: {solution.status}).")

    if not all_pass:
        print("\nWARNING: at least one rule FAILED validation - do not use this roster.")
        return 1
    print("\nAll rules PASS.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
